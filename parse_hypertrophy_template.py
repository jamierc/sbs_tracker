#!/usr/bin/env python3
"""
Parse SBS Hypertrophy Template Excel file and generate JSON data
"""

import openpyxl
import json
import sys
from collections import defaultdict

def clean_value(val):
    """Clean cell values"""
    if val is None:
        return None
    if isinstance(val, str):
        return val.strip()
    return val

def is_tm_row(exercise_name):
    """Check if row is a Training Max row (to be filtered out)"""
    if not exercise_name:
        return False
    if isinstance(exercise_name, str):
        # Check for TM indicators
        if ' TM' in exercise_name or 'TM ' in exercise_name or exercise_name.endswith('TM'):
            return True
    return False

def parse_sheet(sheet, program_name):
    """Parse a single sheet (one program variant like 2x, 3x, etc.)"""
    print(f"Parsing sheet: {sheet.title} for program: {program_name}")

    weeks_data = defaultdict(lambda: defaultdict(list))

    max_row = sheet.max_row
    max_col = sheet.max_column

    # Find week headers in row 1
    week_start_cols = {}
    for col in range(1, max_col + 1):
        cell_val = clean_value(sheet.cell(1, col).value)
        if cell_val and 'Week' in str(cell_val):
            try:
                week_num = int(str(cell_val).split()[-1])
                week_start_cols[week_num] = col
            except (ValueError, IndexError):
                pass

    if not week_start_cols:
        print(f"  No week columns found in {sheet.title}")
        return {}

    print(f"  Found {len(week_start_cols)} weeks")

    # Sort weeks
    sorted_weeks = sorted(week_start_cols.items())

    # For each week, identify which columns contain which data
    # Row 2 has headers: Weight, Reps per normal set, Rep out target, Set goal, Reps on last set, Video, Notes
    week_col_mapping = {}
    for week_num, start_col in sorted_weeks:
        # Find the next week's start column (or max_col)
        next_week_start = max_col + 1
        for next_week, next_col in sorted_weeks:
            if next_week > week_num:
                next_week_start = next_col
                break

        # Map data columns for this week
        week_col_mapping[week_num] = {
            'weight': start_col,
            'reps': start_col + 1,
            'target': start_col + 2,
            'sets': start_col + 3
        }

    # Now parse exercise data
    current_day = None

    for row in range(3, max_row + 1):
        # Get exercise name from column 1
        exercise_name = clean_value(sheet.cell(row, 1).value)

        if not exercise_name:
            continue

        # Check if it's a day header
        if isinstance(exercise_name, str) and 'Day' in exercise_name:
            try:
                current_day = int(exercise_name.split()[-1])
                continue
            except (ValueError, IndexError):
                pass

        # Skip TM rows
        if is_tm_row(exercise_name):
            continue

        # Skip accessory headers
        if isinstance(exercise_name, str) and 'Accessories' in exercise_name:
            continue

        # If we have a current day, this is exercise data
        if current_day and isinstance(exercise_name, str):
            # Parse data for each week
            for week_num, cols in week_col_mapping.items():
                try:
                    weight = clean_value(sheet.cell(row, cols['weight']).value)
                    reps = clean_value(sheet.cell(row, cols['reps']).value)
                    target = clean_value(sheet.cell(row, cols['target']).value)
                    sets = clean_value(sheet.cell(row, cols['sets']).value)

                    # Handle "n/a" or missing target values (deload weeks)
                    if target == 'n/a' or target == 'N/A':
                        target = 0

                    # Only add if we have valid data for weight, reps, and sets
                    if (isinstance(weight, (int, float)) and
                        isinstance(reps, (int, float)) and
                        isinstance(sets, (int, float))):

                        # Convert target to int, default to 0 if not numeric
                        try:
                            target_int = int(target) if isinstance(target, (int, float)) else 0
                        except (ValueError, TypeError):
                            target_int = 0

                        exercise_data = {
                            'name': exercise_name,
                            'weight': float(weight),
                            'reps': int(reps),
                            'target': target_int,
                            'sets': int(sets)
                        }

                        weeks_data[str(week_num)][str(current_day)].append(exercise_data)
                except (ValueError, TypeError) as e:
                    pass

    return dict(weeks_data)

def parse_hypertrophy_template(excel_file):
    """Parse the full hypertrophy template Excel file"""
    print(f"Loading workbook: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    all_programs = {}

    # Map sheet names to program variants
    program_mapping = {
        '2x': '2x',
        '3x': '3x',
        '4x': '4x',
        '5x': '5x',
        '6x': '6x'
    }

    # Parse each sheet
    for sheet in wb.worksheets:
        sheet_name = sheet.title
        print(f"\nChecking sheet: {sheet_name}")

        # Try to match sheet name to program
        program_name = None
        for key, value in program_mapping.items():
            if key.lower() == sheet_name.lower():
                program_name = value
                break

        if program_name:
            weeks_data = parse_sheet(sheet, program_name)
            if weeks_data:
                all_programs[program_name] = weeks_data
                print(f"  ✓ Parsed {program_name}: {len(weeks_data)} weeks")

                # Show sample data
                if '1' in weeks_data and '1' in weeks_data['1']:
                    first_exercises = weeks_data['1']['1']
                    if first_exercises:
                        print(f"    Sample: Week 1, Day 1, First exercise: {first_exercises[0]}")
            else:
                print(f"  ✗ No data extracted for {program_name}")
        else:
            print(f"  Skipping sheet: {sheet_name}")

    return all_programs

def main():
    excel_file = 'SBS Hypertrophy Template.xlsx'
    output_file = 'hypertrophy_programs.json'

    print("=" * 60)
    print("SBS Hypertrophy Template Parser")
    print("=" * 60)

    try:
        all_programs = parse_hypertrophy_template(excel_file)

        if not all_programs:
            print("\n❌ ERROR: No program data was extracted!")
            print("The Excel file structure may be different than expected.")
            sys.exit(1)

        # Save to JSON
        with open(output_file, 'w') as f:
            json.dump(all_programs, f, indent=2)

        print("\n" + "=" * 60)
        print(f"✓ Successfully created {output_file}")
        print("=" * 60)
        print(f"\nPrograms parsed: {', '.join(all_programs.keys())}")

        # Print summary
        for program, weeks in all_programs.items():
            print(f"\n{program} Program:")
            print(f"  - Weeks: {len(weeks)}")
            if weeks:
                first_week = list(weeks.values())[0]
                print(f"  - Days per week: {len(first_week)}")
                if first_week:
                    first_day = list(first_week.values())[0]
                    print(f"  - Exercises on day 1: {len(first_day)}")

    except FileNotFoundError:
        print(f"\n❌ ERROR: File '{excel_file}' not found!")
        print("Make sure the Excel file is in the current directory.")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
