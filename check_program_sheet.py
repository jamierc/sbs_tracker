#!/usr/bin/env python3
"""Check auxiliary exercises in program sheets"""

import openpyxl

excel_file = 'SBS Hypertrophy Template.xlsx'
wb = openpyxl.load_workbook(excel_file, data_only=False)

# Check the 2x sheet
sheet = wb['2x']

print("=" * 60)
print("Checking '2x' sheet for all exercises")
print("=" * 60)

# Look at column A for all exercise names
print("\nAll exercises in column A (first 100 rows):")
for row in range(1, min(100, sheet.max_row + 1)):
    cell_value = sheet.cell(row, 1).value
    if cell_value:
        print(f"Row {row}: {cell_value}")

# Now check for data validations
print("\n" + "=" * 60)
print("Data validations in 2x sheet:")
print("=" * 60)

if hasattr(sheet, 'data_validations'):
    for dv in sheet.data_validations.dataValidation:
        if dv.formula1:
            print(f"\nValidation for cells: {dv.sqref}")
            print(f"  Formula: {dv.formula1}")
            print(f"  Type: {dv.type}")
