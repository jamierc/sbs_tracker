#!/usr/bin/env python3
"""Check data validations in Quick Setup sheet"""

import openpyxl

excel_file = 'SBS Hypertrophy Template.xlsx'
wb = openpyxl.load_workbook(excel_file, data_only=False)

quick_setup = wb['Quick Setup']

print("=" * 60)
print("Quick Setup sheet - Data validations")
print("=" * 60)

if hasattr(quick_setup, 'data_validations'):
    for idx, dv in enumerate(quick_setup.data_validations.dataValidation, 1):
        if dv.formula1:
            print(f"\n#{idx} Validation for cells: {dv.sqref}")
            print(f"  Formula: {dv.formula1}")
            print(f"  Type: {dv.type}")

            # Check which row this is on to understand context
            try:
                cell_ref = str(dv.sqref).split()[0].split(':')[0]
                cell = quick_setup[cell_ref]
                print(f"  Cell: {cell_ref} (row {cell.row}, col {cell.column})")

                # Get the label from column B (which has the auxiliary name)
                label_cell = quick_setup.cell(cell.row, 2)  # Column B
                print(f"  Label (col B): {label_cell.value}")

                # Get the value from column C (the exercise name)
                value_cell = quick_setup.cell(cell.row, 3)  # Column C
                print(f"  Current value (col C): {value_cell.value}")
            except Exception as e:
                print(f"  Error getting context: {e}")

            # If the formula references a range, let's see what's in it
            if dv.formula1 and '!' in dv.formula1:
                try:
                    # Parse the formula to get sheet and range
                    parts = dv.formula1.replace("'", "").split('!')
                    if len(parts) == 2:
                        ref_sheet_name = parts[0]
                        ref_range = parts[1].replace('$', '')

                        print(f"  References: {ref_sheet_name} range {ref_range}")

                        # Get the actual values
                        ref_sheet = wb[ref_sheet_name]

                        # Try to parse the range
                        if ':' in ref_range:
                            start, end = ref_range.split(':')
                            # This is complex, just note it
                            print(f"    (Range from {start} to {end})")
                        else:
                            cell = ref_sheet[ref_range]
                            print(f"    Value: {cell.value}")
                except Exception as e:
                    print(f"  Error parsing formula: {e}")
