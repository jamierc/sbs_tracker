#!/usr/bin/env python3
"""Check auxiliary exercise options in SBS Template"""

import openpyxl

excel_file = 'SBS Hypertrophy Template.xlsx'
wb = openpyxl.load_workbook(excel_file, data_only=False)

# Check all sheets
print("=" * 60)
print("All sheets in workbook:")
for sheet in wb.worksheets:
    print(f"  - {sheet.title}")

print("\n" + "=" * 60)
print("Checking 'Setup' sheet for auxiliary exercises...")
print("=" * 60)

setup_sheet = wb['Setup']

# Look for rows containing "auxiliary" or "Auxiliary"
for row_idx in range(1, min(200, setup_sheet.max_row + 1)):
    for col_idx in range(1, min(20, setup_sheet.max_column + 1)):
        cell = setup_sheet.cell(row_idx, col_idx)
        if cell.value and isinstance(cell.value, str):
            if 'auxiliary' in cell.value.lower() or 'aux' in cell.value.lower():
                print(f"\nRow {row_idx}, Col {col_idx}: {cell.value}")

                # Check if cell has data validation (dropdown)
                if hasattr(cell, 'data_validation') or (setup_sheet.data_validations and
                    any(dv for dv in setup_sheet.data_validations.dataValidation
                        if cell.coordinate in dv.cells)):
                    print(f"  ^ This cell has data validation!")

# Also check for data validations
print("\n" + "=" * 60)
print("Checking data validations in Setup sheet...")
print("=" * 60)

if hasattr(setup_sheet, 'data_validations'):
    for dv in setup_sheet.data_validations.dataValidation:
        if dv.formula1:
            print(f"\nValidation for cells: {dv.sqref}")
            print(f"  Formula: {dv.formula1}")
            print(f"  Type: {dv.type}")

            # Check what cells are affected
            for cell_range in str(dv.sqref).split():
                print(f"  Affected range: {cell_range}")
                # Get a sample cell to see what exercise it's for
                try:
                    first_cell = setup_sheet[cell_range.split(':')[0]]
                    row = first_cell.row
                    # Look at nearby cells to see exercise context
                    for check_col in range(max(1, first_cell.column - 3), first_cell.column):
                        context_cell = setup_sheet.cell(row, check_col)
                        if context_cell.value:
                            print(f"    Context (col {check_col}): {context_cell.value}")
                except:
                    pass
